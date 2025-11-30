"""Модуль для чтения данных из Excel файлов."""
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelReader:
    """Класс для чтения и парсинга Excel файлов."""
    
    # Столбцы по умолчанию согласно ТЗ
    DEFAULT_COLUMNS = {
        "position": 0,      # A - № п.п.
        "supplier": 19,     # T - Наименование поставщика
        "kpp": 21,          # V - КПП
        "inn": 22,          # W - ИНН
        "hyperlink": 23     # X - Гиперссылка
    }
    
    # Данные начинаются со строки 11 (индекс 10)
    DATA_START_ROW = 11
    
    def __init__(self, file_path: str):
        """Инициализация reader'а."""
        self.file_path = Path(file_path)
        self.workbook = None
        self.sheet: Optional[Worksheet] = None
        self.data: Dict[str, Dict] = {}
        self._loaded = False
    
    def load(self) -> Tuple[bool, str]:
        """Загрузка Excel файла."""
        try:
            if not self.file_path.exists():
                return False, f"Файл не найден: {self.file_path}"
            
            self.workbook = load_workbook(str(self.file_path), read_only=True, data_only=True)
            self.sheet = self.workbook.active
            self._loaded = True
            return True, "OK"
        except Exception as e:
            return False, f"Ошибка загрузки Excel: {str(e)}"
    
    def parse(self, column_mapping: Optional[Dict[str, int]] = None) -> Tuple[bool, str]:
        """Парсинг данных из Excel."""
        if not self._loaded:
            return False, "Excel файл не загружен"
        
        columns = column_mapping or self.DEFAULT_COLUMNS
        
        try:
            row_num = self.DATA_START_ROW
            while True:
                # Получаем значение позиции (столбец A)
                position_cell = self.sheet.cell(row=row_num, column=columns["position"] + 1)
                position_value = position_cell.value
                
                if position_value is None or str(position_value).strip() == "":
                    break
                
                # Нормализуем позицию (1.1 -> 1-1)
                position_key = self._normalize_position(str(position_value))
                
                # Читаем данные для этой позиции
                row_data = {}
                for field_name, col_index in columns.items():
                    if field_name == "position":
                        continue
                    cell_value = self.sheet.cell(row=row_num, column=col_index + 1).value
                    row_data[field_name] = str(cell_value) if cell_value is not None else ""
                
                self.data[position_key] = row_data
                row_num += 1
            
            return True, f"Загружено {len(self.data)} записей"
        except Exception as e:
            return False, f"Ошибка парсинга: {str(e)}"
    
    def _normalize_position(self, position: str) -> str:
        """Нормализация позиции: 1.1 -> 1-1"""
        # Убираем лишние пробелы
        position = position.strip()
        # Заменяем точки на дефисы
        position = position.replace(".", "-")
        return position
    
    def get_data_for_file(self, filename: str) -> Optional[Dict]:
        """Получение данных для конкретного файла."""
        # Извлекаем позицию из имени файла (1-1.png -> 1-1)
        position_key = Path(filename).stem
        return self.data.get(position_key)
    
    def get_all_positions(self) -> List[str]:
        """Получение списка всех позиций."""
        return list(self.data.keys())
    
    def get_record_count(self) -> int:
        """Получение количества записей."""
        return len(self.data)
    
    def close(self) -> None:
        """Закрытие файла."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self._loaded = False
    
    def __enter__(self):
        """Context manager вход."""
        self.load()
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager выход."""
        self.close()
        return False


def get_column_letter(index: int) -> str:
    """Преобразование индекса столбца в букву (0 -> A, 1 -> B, ...)."""
    result = ""
    while index >= 0:
        result = chr(index % 26 + ord('A')) + result
        index = index // 26 - 1
    return result


def get_column_index(letter: str) -> int:
    """Преобразование буквы столбца в индекс (A -> 0, B -> 1, ...)."""
    result = 0
    for char in letter.upper():
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result - 1
